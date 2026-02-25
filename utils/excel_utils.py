"""
ExcelUtils — Daily attendance Excel file management.
One file per day: attendance_records/YYYY-MM-DD.xlsx
Three sheets: Students, Staff, Admin
"""
import os
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BASE_DIR         = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RECORDS_DIR      = os.path.join(BASE_DIR, "attendance_records")

COLUMNS = ["Name", "ID", "Course/Dept", "Time", "Date", "Status", "Image Path"]

SHEET_MAP = {
    "student": "Students",
    "staff":   "Staff",
    "admin":   "Admin",
}

# Header fill colours per sheet
HEADER_FILLS = {
    "Students": "1E88E5",   # blue
    "Staff":    "43A047",   # green
    "Admin":    "6D4C41",   # brown
}


def _get_file_path(date_str: str) -> str:
    os.makedirs(RECORDS_DIR, exist_ok=True)
    return os.path.join(RECORDS_DIR, f"{date_str}.xlsx")


def _thin_border() -> Border:
    thin = Side(border_style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header(ws, fill_hex: str):
    """Style the header row of a sheet."""
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.row_dimensions[1].height = 22


def _get_or_create_workbook(date_str: str) -> openpyxl.Workbook:
    """Load existing workbook or create a new one with 3 sheets."""
    path = _get_file_path(date_str)
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path)
            return wb
        except Exception:
            pass  # corrupt → recreate

    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default Sheet
    for sheet_name in ["Students", "Staff", "Admin"]:
        ws = wb.create_sheet(sheet_name)
        _apply_header(ws, HEADER_FILLS[sheet_name])
    return wb


def append_to_excel(date_str: str, record: Dict, role: str) -> bool:
    """
    Append a single attendance record to the correct sheet.
    record keys: name, user_id, course, department, time, date, status, image_path
    Returns True on success.
    """
    sheet_name = SHEET_MAP.get(role.lower(), "Students")
    path = _get_file_path(date_str)

    try:
        wb = _get_or_create_workbook(date_str)
        ws = wb[sheet_name]

        row = [
            record.get("name", ""),
            record.get("user_id", ""),
            f"{record.get('course', '')} / {record.get('department', '')}",
            record.get("time", ""),
            record.get("date", ""),
            record.get("status", "present"),
            record.get("image_path", ""),
        ]

        ws.append(row)

        # Style the new data row
        data_row_idx = ws.max_row
        alt_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        ) if data_row_idx % 2 == 0 else None

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="center")
            cell.border = _thin_border()
            if alt_fill:
                cell.fill = alt_fill

        wb.save(path)
        return True
    except Exception as e:
        from utils.logger import logger
        logger.error(f"Excel write error for {date_str}: {e}")
        return False


def get_daily_file_path(date_str: Optional[str] = None) -> str:
    """Return the path for today's (or given date) Excel file."""
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    return _get_file_path(date_str)


def export_date_range_to_excel(
    start_date: str, end_date: str, output_path: str
) -> bool:
    """
    Merge multiple daily files into one export workbook.
    Used by admin export feature.
    """
    from database.db_manager import DatabaseManager
    db = DatabaseManager()

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for sheet_name, role in [("Students", "student"),
                                   ("Staff", "staff"),
                                   ("Admin", "admin")]:
            ws = wb.create_sheet(sheet_name)
            _apply_header(ws, HEADER_FILLS[sheet_name])

            # Fetch from DB for the date range
            from database.db_manager import DatabaseManager
            db = DatabaseManager()
            import sqlite3
            with db._main_conn() as conn:
                rows = conn.execute(
                    """SELECT name, user_id, course, department, time, date,
                              status, image_path
                       FROM attendance
                       WHERE role=? AND date BETWEEN ? AND ?
                       ORDER BY date, time""",
                    (role, start_date, end_date),
                ).fetchall()

            for r in rows:
                ws.append([
                    r["name"], r["user_id"],
                    f"{r['course']} / {r['department']}",
                    r["time"], r["date"], r["status"], r["image_path"],
                ])

        wb.save(output_path)
        return True
    except Exception as e:
        from utils.logger import logger
        logger.error(f"Export error: {e}")
        return False
